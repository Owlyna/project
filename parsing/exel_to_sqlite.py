import os
import sqlite3
import openpyxl


def export_to_sqlite():
    '''Экспорт данных из xlsx в sqlite'''
    # 1. Создание и подключение к базе

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Имя базы
    base_name = 'db.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect(prj_dir + '/' + base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    cursor.execute('CREATE TABLE IF NOT EXISTS films (id INTEGER PRIMARY KEY AUTOINCREMENT, НазваниеФильма TEXT NOT NULL, ГодВыпуска INTEGER NOT NULL, СтранаВыпуска TEXT NOT NULL, Жанр TEXT NOT NULL, Режиссер TEXT NOT NULL, Оценка REAL NOT NULL, КоличествоОтзывов INTEGER NOT NULL)')
    #cursor.execute('CREATE TABLE IF NOT EXISTS films (idФильма INTEGER PRIMARY KEY AUTOINCREMENT, НазваниеФильма TEXT, ГодВыпуска INTEGER, idСтраныВыпуска INT, Жанр TEXT, Режиссер TEXT, Оценка REAL, КоличествоОтзывов INTEGER)')
    #cursor.execute('CREATE TABLE IF NOT EXISTS countries (idСтаныВыпуска INTEGER PRIMARY KEY AUTOINCREMENT, СтранаВыпуска TEXT)')

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('database.xlsx', data_only=True)
    sheet = file_to_read['Sheet1']

    # Цикл по строкам начиная со второй (в первой заголовки)

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 8 ( 9 не включая)
        for col in range(1, 9):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)

    # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы
        cursor.execute("INSERT INTO films (НазваниеФильма, ГодВыпуска, СтранаВыпуска, Жанр, Режиссер, Оценка, КоличествоОтзывов) VALUES ( ?, ?, ?, ?, ?, ?, ?);", (data[1], data[2], data[3], data[4], data[5], data[6], data[7]))
        #cursor.execute("INSERT INTO films (НазваниеФильма, ГодВыпуска, Жанр, Режиссер, Оценка, КоличествоОтзывов) VALUES ( ?, ?, ?, ?, ?, ?);", (data[0], data[1], data[3], data[4], data[5], data[6]))
        #cursor.execute("INSERT INTO countries (СтранаВыпуска) VALUES (?);", (data[2]))


    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


def clear_base():
    '''Очистка базы sqlite'''

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 'db.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute("DELETE FROM db")
    connect.commit()
    connect.close()


# Запуск функции
export_to_sqlite()


